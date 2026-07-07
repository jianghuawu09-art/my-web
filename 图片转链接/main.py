import os
import uuid
import requests
import time
import sqlite3
import json
import io
import tempfile
from flask import Flask, request, jsonify, render_template_string, send_file
from flask_cors import CORS
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# ==================== 薄荷图床配置 ====================
MINT_UPLOAD_URL = 'https://ok.a2k6.com/yixing01/api/upload/'
MINT_API_TOKEN = '18fe510915103f9e2ca3'

# ==================== 钉钉配置 ====================
DT_CLIENT_ID = 'dingyzwmwrqoumv2xyjb'
DT_CLIENT_SECRET = 'mh0WB_3cqnOI3znHRr7Yj1BD5MQjwr_rrFHTYZtaJAx-4J2oQavuRk-KwU8kddH3'
BASE_ID = '4lgGw3P8vRLPeRmXsp40mBj285daZ90D'
SHEET_ID = 'hERWDMS'
OPERATOR_ID = "fADzjBkYl9VEvxiPwAKrF2QiEiE"

SAVE_DIR = os.path.join(os.path.dirname(__file__), 'temp_images')
DB_PATH = os.path.join(os.path.dirname(__file__), 'upload_records.db')
os.makedirs(SAVE_DIR, exist_ok=True)

app = Flask(__name__)
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# Excel 工作簿存储
excel_workbooks = {}


@app.before_request
def log_request():
    print(f"\n🔍 请求到达: {request.method} {request.path}")
    print(f"  Content-Type: {request.content_type}")
    if request.files:
        print(f"  Files: {list(request.files.keys())}")


def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS upload_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL,
            image_url TEXT,
            result_url TEXT,
            status TEXT NOT NULL,
            error_msg TEXT,
            created_at TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()


def save_record(filename, image_url, result_url, status, error_msg=None):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO upload_records (filename, image_url, result_url, status, error_msg, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (filename, image_url or '', result_url or '', status, error_msg or '', datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit()
    conn.close()


def get_records(page=1, page_size=10):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM upload_records')
    total = cursor.fetchone()[0]
    offset = (page - 1) * page_size
    cursor.execute('''
        SELECT id, filename, image_url, result_url, status, error_msg, created_at
        FROM upload_records
        ORDER BY id DESC
        LIMIT ? OFFSET ?
    ''', (page_size, offset))
    records = []
    for row in cursor.fetchall():
        records.append({
            'id': row[0],
            'filename': row[1],
            'image_url': row[2],
            'result_url': row[3],
            'status': row[4],
            'error_msg': row[5],
            'created_at': row[6]
        })
    conn.close()
    return {
        'total': total,
        'page': page,
        'page_size': page_size,
        'records': records
    }


init_db()


# ==================== 薄荷图床上传 ====================
def upload_to_mint(local_path, filename):
    file_handle = None
    try:
        file_handle = open(local_path, 'rb')
        files = {
            'uploadedFile': (filename, file_handle, 'image/jpeg')
        }
        data = {
            'api_token': MINT_API_TOKEN,
            'upload_format': 'file',
            'mode': '1',
            'uploadPath': "",
            'watermark': '0',
        }
        headers = {"Accept-Encoding": "identity"}
        response = requests.post(
            MINT_UPLOAD_URL,
            data=data,
            files=files,
            headers=headers,
            timeout=25
        )
        raw_text = response.text.strip()
        print("薄荷接口原始返回：", repr(raw_text))
        if not raw_text:
            print("❌ 薄荷返回空白，上传失败")
            return None
        res_json = response.json()
        if res_json.get("status") == "success":
            img_url = res_json.get("url")
            print(f"✅ 薄荷上传成功，链接：{img_url}")
            return img_url
        else:
            print(f"❌ 薄荷业务报错：{res_json}")
            return None
    except Exception as e:
        print(f"❌ 薄荷上传异常：{e}")
        return None
    finally:
        if file_handle:
            file_handle.close()


# ==================== 钉钉 API 函数 ====================
def get_dingtalk_token():
    url = "https://api.dingtalk.com/v1.0/oauth2/accessToken"
    res = requests.post(url, json={
        "appKey": DT_CLIENT_ID,
        "appSecret": DT_CLIENT_SECRET
    })
    return res.json()["accessToken"]


def get_all_records(token):
    url = f"https://api.dingtalk.com/v1.0/notable/bases/{BASE_ID}/sheets/{SHEET_ID}/records/list?operatorId={OPERATOR_ID}"
    headers = {
        "x-acs-dingtalk-access-token": token,
        "Content-Type": "application/json"
    }
    res = requests.post(url, headers=headers, json={"pageSize": 500})
    return res.json()


def update_transfer_link(token, record_id, final_link):
    url = f"https://api.dingtalk.com/v1.0/notable/bases/{BASE_ID}/sheets/{SHEET_ID}/records?operatorId={OPERATOR_ID}"
    headers = {
        "x-acs-dingtalk-access-token": token,
        "Content-Type": "application/json"
    }
    payload = {
        "records": [{
            "id": record_id,
            "fields": {
                "转链接": {
                    "link": final_link,
                    "text": ""
                }
            }
        }]
    }
    try:
        resp = requests.put(url, json=payload, headers=headers)
        print(f"✅ 写入成功 行:{record_id} | 状态码:{resp.status_code}")
        return True
    except Exception as e:
        print(f"❌ 写入失败：{e}")
        return False


def process_single_record(token, record):
    record_id = record["id"]
    fields = record.get("fields", {})
    images = fields.get("产品图", [])
    if not images:
        print(f"⚠️ 行 {record_id} 没有产品图，跳过")
        return {"record_id": record_id, "links": [], "success": False, "msg": "无产品图"}
    print(f"\n🖼️ 处理行 {record_id}，共 {len(images)} 张图片")
    mint_urls = []
    for img in images:
        filename = img.get("filename")
        img_url = img.get("url")
        if not filename or not img_url:
            continue
        local_path = os.path.join(SAVE_DIR, filename)
        try:
            img_data = requests.get(img_url, timeout=15).content
            with open(local_path, "wb") as f:
                f.write(img_data)
            print(f"  下载完成：{filename}")
        except Exception as e:
            print(f"  图片下载失败：{e}")
            continue
        mint_url = upload_to_mint(local_path, filename)
        if mint_url:
            mint_urls.append(mint_url)
        if os.path.exists(local_path):
            try:
                os.remove(local_path)
            except:
                pass
    if mint_urls:
        final_link = "\n".join(mint_urls)
        success = update_transfer_link(token, record_id, final_link)
        return {"record_id": record_id, "links": mint_urls, "success": success}
    return {"record_id": record_id, "links": [], "success": False, "msg": "图片处理失败"}


# ==================== Flask 接口 ====================
@app.route('/api/upload', methods=['POST'])
def upload_images():
    if 'files' not in request.files:
        return jsonify({"code": 400, "msg": "没有上传文件", "data": []})

    files = request.files.getlist('files')
    results = []

    print(f"\n📥 收到 {len(files)} 个文件")

    for i, file in enumerate(files):
        original_filename = file.filename if file.filename else f"unknown_{i+1}"
        content_type = file.content_type

        print(f"  文件 {i+1}: filename={repr(original_filename)}, content_type={content_type}")

        if original_filename == '' or not original_filename:
            continue

        if not content_type or not content_type.startswith('image/'):
            results.append({
                "filename": original_filename,
                "success": False,
                "error": "不是有效的图片文件"
            })
            save_record(original_filename, '', '', 'failed', '不是有效的图片文件')
            continue

        file_ext = os.path.splitext(original_filename)[1] or '.jpg'
        unique_name = f"{uuid.uuid4().hex}{file_ext}"
        local_path = os.path.join(SAVE_DIR, unique_name)

        try:
            content = file.read()
            with open(local_path, 'wb') as f:
                f.write(content)

            time.sleep(0.5)

            img_url = upload_to_mint(local_path, unique_name)

            if img_url:
                results.append({
                    "filename": original_filename,
                    "success": True,
                    "url": img_url
                })
                save_record(original_filename, img_url, img_url, 'success')
            else:
                results.append({
                    "filename": original_filename,
                    "success": False,
                    "error": "上传失败"
                })
                save_record(original_filename, '', '', 'failed', '上传失败')
        except Exception as e:
            print(f"  ❌ 处理异常：{e}")
            results.append({
                "filename": original_filename,
                "success": False,
                "error": str(e)
            })
            save_record(original_filename, '', '', 'failed', str(e))
        finally:
            if os.path.exists(local_path):
                try:
                    os.remove(local_path)
                except:
                    pass

    return jsonify({"code": 200, "msg": "处理完成", "data": results})


@app.route('/api/records', methods=['GET'])
def get_upload_records():
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 10))
    data = get_records(page, page_size)
    return jsonify({"code": 200, "msg": "查询成功", "data": data})


@app.route('/api/records/<int:record_id>', methods=['DELETE'])
def delete_record(record_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM upload_records WHERE id = ?', (record_id,))
    conn.commit()
    affected = cursor.rowcount
    conn.close()
    if affected > 0:
        return jsonify({"code": 200, "msg": "删除成功"})
    else:
        return jsonify({"code": 404, "msg": "记录不存在"})


# ==================== Excel 链接填写接口 ====================
@app.route('/api/excel/upload', methods=['POST'])
def upload_excel():
    print("\n📥 Excel上传请求到达")
    print(f"  请求方法: {request.method}")
    print(f"  请求路径: {request.path}")
    print(f"  Content-Type: {request.content_type}")
    print(f"  文件列表: {list(request.files.keys())}")
    
    if 'file' not in request.files:
        print("❌ 未找到文件")
        return jsonify({'error': '未找到文件'}), 400
    
    file = request.files['file']
    print(f"  文件名: {file.filename}")
    print(f"  文件大小: {len(file.read())} bytes")
    file.seek(0)
    
    if file.filename == '':
        print("❌ 未选择文件")
        return jsonify({'error': '未选择文件'}), 400
    
    if not file.filename.endswith(('.xlsm', '.xlsx', '.xls')):
        print(f"❌ 文件格式不支持: {file.filename}")
        return jsonify({'error': '文件格式不支持'}), 400
    
    file_id = str(uuid.uuid4())
    file_path = os.path.join(tempfile.gettempdir(), f'{file_id}.xlsm')
    file.save(file_path)
    print(f"  文件已保存到: {file_path}")
    
    try:
        wb = load_workbook(file_path, keep_vba=True)
        ws = wb.active
        print(f"  工作表名称: {ws.title}")
        print(f"  最大行数: {ws.max_row}, 最大列数: {ws.max_column}")
        
        sku_list = []
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value is not None:
                sku_list.append({
                    'value': str(cell_value).strip(),
                    'rowIndex': row - 1
                })
        print(f"  提取SKU数量: {len(sku_list)}")
        
        # 读取第4行，查找"Main Image URL"列
        main_image_url_col = None
        main_image_url_col_letter = 'T'  # 默认从T列开始
        
        # 获取单元格文本内容的辅助函数（处理富文本）
        def get_cell_text(cell):
            val = cell.value
            if val is None:
                return ''
            if isinstance(val, str):
                return val
            # 处理富文本格式
            try:
                if hasattr(val, 'text'):
                    return str(val.text)
                # 如果是列表（openpyxl富文本格式）
                if isinstance(val, list):
                    texts = []
                    for item in val:
                        if hasattr(item, 'text'):
                            texts.append(str(item.text))
                        elif isinstance(item, str):
                            texts.append(item)
                    return ''.join(texts)
            except:
                pass
            return str(val)
        
        # 检查合并单元格
        print(f"  合并单元格数量: {len(ws.merged_cells.ranges)}")
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= 4 <= merged_range.max_row:
                print(f"  第4行在合并区域内: {merged_range.coord} (min_col={merged_range.min_col}, max_col={merged_range.max_col})")
        
        if ws.max_row >= 4:
            print(f"  正在第4行查找 'Main Image URL'...")
            # 扩大列遍历范围到100列，确保不会遗漏
            max_col_to_search = max(ws.max_column, 100)
            print(f"  最大列数: {ws.max_column}, 搜索范围: 1-{max_col_to_search}")
            
            for col in range(1, max_col_to_search + 1):
                cell = ws.cell(row=4, column=col)
                cell_value = cell.value
                cell_type = type(cell_value).__name__
                cell_str = get_cell_text(cell).strip()
                cell_str_lower = cell_str.lower()
                col_letter = get_column_letter(col)
                
                # 输出前30列和所有非空列的值，便于调试
                if col <= 30 or cell_str:
                    print(f"    第{col}列 ({col_letter}): type={cell_type}, value={repr(cell_value)}, text={repr(cell_str)}")
                
                # 匹配规则1：精确匹配（不区分大小写）
                if cell_str_lower == 'main image url':
                    main_image_url_col = col
                    main_image_url_col_letter = col_letter
                    print(f"  ✅ 精确匹配到Main Image URL列: {col_letter} (第{col}列)")
                    break
                # 匹配规则2：包含匹配（不区分大小写）
                elif 'main image url' in cell_str_lower:
                    main_image_url_col = col
                    main_image_url_col_letter = col_letter
                    print(f"  ✅ 模糊匹配到Main Image URL列: {col_letter} (第{col}列), 值: {cell_str}")
                    break
                # 匹配规则3：只匹配 "Main Image"（不区分大小写）
                elif 'main image' in cell_str_lower and 'url' in cell_str_lower:
                    main_image_url_col = col
                    main_image_url_col_letter = col_letter
                    print(f"  ✅ 拆分匹配到Main Image URL列: {col_letter} (第{col}列), 值: {cell_str}")
                    break
            
            if main_image_url_col is None:
                print(f"  ⚠️ 未在第4行找到 'Main Image URL'，使用默认列 T")
        else:
            print(f"  ⚠️ 表格行数不足4行 (max_row={ws.max_row})，使用默认列 T")
        
        excel_workbooks[file_id] = {
            'workbook': wb,
            'worksheet': ws,
            'file_path': file_path,
            'filename': file.filename,
            'sku_list': sku_list,
            'main_image_url_col': main_image_url_col,
            'main_image_url_col_letter': main_image_url_col_letter
        }
        
        print(f"✅ Excel上传成功: {file.filename}")
        print(f"  存储的Main Image URL列: {main_image_url_col_letter} (第{main_image_url_col}列)")
        print(f"  excel_workbooks keys: {list(excel_workbooks[file_id].keys())}")
        return jsonify({
            'fileId': file_id,
            'filename': file.filename,
            'skuCount': len(sku_list),
            'mainImageUrlCol': main_image_url_col_letter
        })
    
    except Exception as e:
        print(f"❌ 文件解析失败: {str(e)}")
        if os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({'error': f'文件解析失败: {str(e)}'}), 500


@app.route('/api/excel/sku/search', methods=['GET'])
def search_sku():
    file_id = request.args.get('fileId')
    query = request.args.get('query', '').strip().lower()
    
    if not file_id or file_id not in excel_workbooks:
        return jsonify({'error': '文件不存在'}), 404
    
    sku_list = excel_workbooks[file_id]['sku_list']
    
    if not query:
        return jsonify({'matches': []})
    
    matches = [
        item for item in sku_list 
        if query in item['value'].lower()
    ][:10]
    
    return jsonify({'matches': matches})


@app.route('/api/excel/row/select', methods=['POST'])
def select_row():
    data = request.get_json()
    file_id = data.get('fileId')
    row_index = data.get('rowIndex')
    
    if not file_id or file_id not in excel_workbooks:
        return jsonify({'error': '文件不存在'}), 404
    
    wb_data = excel_workbooks[file_id]
    ws = wb_data['worksheet']
    
    row_num = row_index + 1
    row_data = []
    
    for col in range(1, min(ws.max_column + 1, 30)):
        cell = ws.cell(row=row_num, column=col)
        col_letter = get_column_letter(col)
        header_cell = ws.cell(row=1, column=col)
        
        row_data.append({
            'col': col_letter,
            'value': str(cell.value) if cell.value is not None else '',
            'header': str(header_cell.value) if header_cell.value is not None else ''
        })
    
    return jsonify({
        'rowIndex': row_index,
        'rowNum': row_num,
        'skuValue': str(ws.cell(row=row_num, column=1).value or '').strip(),
        'rowData': row_data
    })


@app.route('/api/excel/links/write', methods=['POST'])
def write_links():
    data = request.get_json()
    file_id = data.get('fileId')
    row_index = data.get('rowIndex')
    links = data.get('links', [])
    
    if not file_id or file_id not in excel_workbooks:
        return jsonify({'error': '文件不存在'}), 404
    
    if row_index is None or not links:
        return jsonify({'error': '参数不完整'}), 400
    
    wb_data = excel_workbooks[file_id]
    ws = wb_data['worksheet']
    
    # 使用存储的Main Image URL列作为起始列，不依赖前端参数
    start_col = wb_data.get('main_image_url_col_letter', 'T')
    print(f"  写入链接：起始列={start_col}，行={row_index + 1}")
    
    row_num = row_index + 1
    start_col_idx = column_index_from_string(start_col)
    
    written_cells = []
    
    for i, link in enumerate(links):
        if not link.strip():
            continue
        col_idx = start_col_idx + i
        col_letter = get_column_letter(col_idx)
        ws.cell(row=row_num, column=col_idx, value=link.strip())
        written_cells.append({
            'row': row_num,
            'col': col_letter,
            'value': link.strip()
        })
    
    row_data = []
    for col in range(1, min(ws.max_column + 1, 30)):
        cell = ws.cell(row=row_num, column=col)
        col_letter = get_column_letter(col)
        header_cell = ws.cell(row=1, column=col)
        
        row_data.append({
            'col': col_letter,
            'value': str(cell.value) if cell.value is not None else '',
            'header': str(header_cell.value) if header_cell.value is not None else '',
            'modified': any(c['row'] == row_num and c['col'] == col_letter for c in written_cells)
        })
    
    return jsonify({
        'success': True,
        'writtenCount': len(written_cells),
        'rowData': row_data,
        'writtenCells': written_cells
    })


@app.route('/api/excel/export', methods=['GET'])
def export_excel():
    file_id = request.args.get('fileId')
    
    if not file_id or file_id not in excel_workbooks:
        return jsonify({'error': '文件不存在'}), 404
    
    wb_data = excel_workbooks[file_id]
    wb = wb_data['workbook']
    original_filename = wb_data['filename']
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    base_name = os.path.splitext(original_filename)[0]
    export_filename = f'{base_name}_已处理.xlsm'
    
    return send_file(
        output,
        as_attachment=True,
        download_name=export_filename,
        mimetype='application/vnd.ms-excel.sheet.macroEnabled.12'
    )


@app.route('/api/excel/preview', methods=['GET'])
def get_preview():
    file_id = request.args.get('fileId')
    row_index = int(request.args.get('rowIndex', 0))
    
    if not file_id or file_id not in excel_workbooks:
        return jsonify({'error': '文件不存在'}), 404
    
    wb_data = excel_workbooks[file_id]
    ws = wb_data['worksheet']
    
    row_num = row_index + 1
    preview_rows = []
    
    start_row = max(1, row_num - 2)
    end_row = min(ws.max_row, row_num + 2)
    
    headers = []
    for col in range(1, min(ws.max_column + 1, 30)):
        col_letter = get_column_letter(col)
        header_val = ws.cell(row=1, column=col).value
        headers.append({
            'col': col_letter,
            'header': str(header_val) if header_val is not None else ''
        })
    
    for r in range(start_row, end_row + 1):
        cells = []
        for col in range(1, min(ws.max_column + 1, 30)):
            cell = ws.cell(row=r, column=col)
            col_letter = get_column_letter(col)
            cells.append({
                'col': col_letter,
                'value': str(cell.value) if cell.value is not None else ''
            })
        preview_rows.append({
            'rowIndex': r - 1,
            'rowNum': r,
            'isMatched': r == row_num,
            'cells': cells
        })
    
    return jsonify({
        'headers': headers,
        'rows': preview_rows
    })


# ==================== 钉钉表格接口 ====================
@app.route('/api/dingtalk/records', methods=['GET'])
def get_dingtalk_records():
    try:
        token = get_dingtalk_token()
        records_data = get_all_records(token)
        records = records_data.get("records", [])
        result = []
        for record in records:
            fields = record.get("fields", {})
            images = fields.get("产品图", [])
            transfer_link = fields.get("转链接", {})
            link_value = ""
            if isinstance(transfer_link, dict):
                link_value = transfer_link.get("link", "")
            elif isinstance(transfer_link, str):
                link_value = transfer_link
            image_list = []
            for img in images:
                image_list.append({
                    "filename": img.get("filename", ""),
                    "url": img.get("url", "")
                })
            result.append({
                "record_id": record.get("id"),
                "title": fields.get("标题", ""),
                "images": image_list,
                "transfer_link": link_value,
                "has_images": len(images) > 0
            })
        return jsonify({"code": 200, "msg": "查询成功", "data": result})
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e)})


@app.route('/api/dingtalk/process/<record_id>', methods=['POST'])
def process_dingtalk_record(record_id):
    try:
        token = get_dingtalk_token()
        records_data = get_all_records(token)
        target_record = None
        for record in records_data.get("records", []):
            if record["id"] == record_id:
                target_record = record
                break
        if not target_record:
            return jsonify({"code": 404, "msg": "记录不存在"})
        result = process_single_record(token, target_record)
        return jsonify({"code": 200, "msg": "处理完成", "data": result})
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e)})


@app.route('/api/dingtalk/process-all', methods=['POST'])
def process_all_dingtalk_records():
    try:
        token = get_dingtalk_token()
        records_data = get_all_records(token)
        records = records_data.get("records", [])
        total_result = []
        for record in records:
            result = process_single_record(token, record)
            total_result.append(result)
        return jsonify({"code": 200, "msg": "执行完成", "data": total_result})
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e)})


@app.route('/')
def index():
    with open(os.path.join(os.path.dirname(__file__), 'index.html'), 'r', encoding='utf-8') as f:
        return f.read()


if __name__ == '__main__':
    print("🚀 图片转链接服务启动中...")
    print(f"📍 本机地址: http://localhost:7777")
    print(f"🌐 局域网地址: http://192.168.1.121:7777")
    app.run(host='0.0.0.0', port=7777, threaded=True)
