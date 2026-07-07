import requests
from openpyxl import Workbook
import io

wb = Workbook()
ws = wb.active
ws['A1'] = 'settings=feedTy'
ws['A2'] = 'YZUS-YXQ65DBS2L'
ws['A3'] = 'YZUS-YXQ65DBS3L'
ws['A4'] = 'YZUS-YXQ65DBS12'

buf = io.BytesIO()
wb.save(buf)
buf.seek(0)

url = 'http://127.0.0.1:7777/api/excel/upload'
files = {'file': ('test.xlsx', buf, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
r = requests.post(url, files=files)
print(f"Status: {r.status_code}")
print(f"Response: {r.text[:1000]}")
